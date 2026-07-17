import os
import re
import json
import requests
import csv
import io
from flask import Flask, request, jsonify
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

UPSTASH_URL = os.environ.get("UPSTASH_REDIS_REST_URL")
UPSTASH_TOKEN = os.environ.get("UPSTASH_REDIS_REST_TOKEN")
SLACK_BOT_TOKEN = os.environ.get("SLACK_BOT_TOKEN")

def redis_set(key, value):
    try:
        headers = {"Authorization": f"Bearer {UPSTASH_TOKEN}"}
        serialized = json.dumps(value)
        url = f"{UPSTASH_URL}/set/{key}"
        response = requests.post(url, headers=headers, json=serialized)
        print(f"Upstash response: {response.status_code} {response.text[:100]}")
    except Exception as e:
        print(f"Upstash error: {e}")

def redis_get(key):
    try:
        headers = {"Authorization": f"Bearer {UPSTASH_TOKEN}"}
        res = requests.get(f"{UPSTASH_URL}/get/{key}", headers=headers)
        result = res.json().get('result')
        if result:
            return json.loads(result) if isinstance(result, str) else result
    except: pass
    return None

def parse_daily_wash(text):
    data = {}
    total = re.search(r'(\d+)\s*Total Routes', text, re.IGNORECASE)
    if total: data['total_routes'] = total.group(1)
    c1 = re.search(r'(\d+)\s*Cycle 1', text, re.IGNORECASE)
    if c1: data['cycle_1'] = c1.group(1)
    c0 = re.search(r'(\d+)\s*Cycle 0', text, re.IGNORECASE)
    if c0: data['cycle_0'] = c0.group(1)
    splits = re.search(r'(\d+)\s*Split[s]?[\s\-]*(.*?)(?:\n|$)', text, re.IGNORECASE)
    if splits:
        data['splits'] = splits.group(1)
        if splits.group(2).strip(): data['split_vans'] = splits.group(2).strip()
    dropped = re.search(r'(\d+)\s*dropped[\s\-]*(.*?)(?:\n|$)', text, re.IGNORECASE)
    if dropped:
        data['dropped'] = dropped.group(1)
        if dropped.group(2).strip(): data['dropped_vans'] = dropped.group(2).strip()
    extras = re.search(r'(\d+)\s*extras?', text, re.IGNORECASE)
    if extras: data['extras'] = extras.group(1)
    sweeper = re.search(r'(\d+)\s*sweeper', text, re.IGNORECASE)
    if sweeper: data['sweeper'] = sweeper.group(1)
    terminations = re.search(r'(\d+)\s*Terminations?', text, re.IGNORECASE)
    if terminations: data['terminations'] = terminations.group(1)
    training = re.search(r'(\d+)\s*Training', text, re.IGNORECASE)
    if training: data['training'] = training.group(1)
    c1_waves = re.findall(r'Wave (\d+)\s*First van in[:\-\s]*([\d:]+).*?Last van out[:\-\s]*([\d:]+)', text, re.IGNORECASE)
    if c1_waves:
        data['c1_waves'] = [{'wave': w[0], 'first_in': w[1], 'last_out': w[2]} for w in c1_waves]
    return data

def parse_evening_wash(text):
    data = {}
    routes = re.search(r'(\d+)\s*Total Routes', text)
    if routes: data['total_routes'] = routes.group(1)
    dropped = re.search(r'(\d+)\s*dropped', text)
    if dropped: data['dropped'] = dropped.group(1)
    rescues = re.search(r'Rescues?:(.*?)(?:\n\n|\Z)', text, re.DOTALL)
    if rescues: data['rescues'] = rescues.group(1).strip()
    terminations = re.search(r'(\d+)\s*Terminations', text)
    if terminations: data['terminations'] = terminations.group(1)
    splits = re.search(r'(\d+)\s*Split', text)
    if splits: data['splits'] = splits.group(1)
    training = re.search(r'(\d+)\s*Training', text)
    if training: data['training'] = training.group(1)
    highlights = re.search(r'Highlights?:(.*?)(?:\n\n|\Z)', text, re.DOTALL | re.IGNORECASE)
    if highlights: data['highlights'] = highlights.group(1).strip()
    return data

def parse_capacity(text):
    weeks = []
    current_week = None
    current_cycle = None
    days = []
    for line in text.split('\n'):
        line = line.strip()
        if not line:
            continue
        week_match = re.search(r'week\s*(\d+).*?(cycle\s*\d+)', line, re.IGNORECASE)
        if week_match:
            if current_week and days:
                weeks.append({'week': current_week, 'cycle': current_cycle, 'days': days})
                days = []
            current_week = week_match.group(1)
            current_cycle = week_match.group(2).strip()
            continue
        cycle_match = re.search(r'(cycle\s*\d+)', line, re.IGNORECASE)
        if cycle_match and not week_match:
            if current_week and days:
                weeks.append({'week': current_week, 'cycle': current_cycle, 'days': days})
                days = []
            current_cycle = cycle_match.group(1).strip()
            continue
        day_match = re.search(r'(sun|mon|tue|wed|thu|fri|sat)\w*[\s\-:]*(\d+)\s*RT[\s\-]*(\d+)\s*ECP[\s\-]*(\d+)\s*DA', line, re.IGNORECASE)
        if day_match and current_week:
            days.append({'day': day_match.group(1).capitalize(), 'rt': day_match.group(2), 'ecp': day_match.group(3), 'das': day_match.group(4)})
    if current_week and days:
        weeks.append({'week': current_week, 'cycle': current_cycle, 'days': days})
    return weeks

def parse_fleet_report(text):
    data = {'raw': text, 'date': text.split('\n')[0]}
    van_ids = re.findall(r'\b([A-Z]{1,3}\d+)\b', text)
    data['vans_mentioned'] = list(set(van_ids))
    grounded = re.findall(r'([A-Z]{1,3}\d+).*?ground', text, re.IGNORECASE)
    data['grounded'] = list(set(grounded))
    ungrounded = re.findall(r'([A-Z]{1,3}\d+).*?unground', text, re.IGNORECASE)
    data['ungrounded'] = list(set(ungrounded))
    afs = re.findall(r'AFS.*?([A-Z]{1,3}\d+)', text, re.IGNORECASE)
    data['afs_mentions'] = list(set(afs))
    pave_setup = re.search(r'PAVE Setup:\s*(\d+)', text, re.IGNORECASE)
    if pave_setup: data['pave_total'] = pave_setup.group(1)
    pave_done = re.findall(r'PAVE Complete:\s*([A-Z]{1,3}\d+)', text, re.IGNORECASE)
    data['pave_completed'] = pave_done
    repairs = re.findall(r'([A-Z]{1,3}\d+).*?repair', text, re.IGNORECASE)
    data['repairs'] = list(set(repairs))
    return data

def parse_incident_report(text):
    data = {'raw': text, 'date': text.split('\n')[0], 'status': 'Open', 'photos': []}
    
    fields = [
        'Driver', 'Type', 'Date/Time', 'Location', 'CRS Case #',
        'LMET #', 'Description', 'Statement', 'LMET Called',
        'Law Enforcement', 'Photos'
    ]
    
    # Build pattern that stops at next field name
    for i, field in enumerate(fields):
        next_fields = fields[i+1:] if i+1 < len(fields) else []
        stop = '|'.join([re.escape(f + ':') for f in next_fields]) if next_fields else '$'
        pattern = re.escape(field) + r':\s*(.+?)(?=' + stop + r'|\Z)'
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            value = match.group(1).strip()
            key = field.lower().replace('/', '_').replace(' ', '_').replace('#', 'number')
            data[key] = value
    
    # Clean up key names
    if 'crs_case_number' in data: data['crs_case'] = data.pop('crs_case_number')
    if 'lmet_number' in data: data['lmet_number'] = data['lmet_number']
    if 'date_time' in data: data['incident_datetime'] = data.pop('date_time')
    if 'lmet_called' in data: data['lmet_called'] = data['lmet_called']
    if 'law_enforcement' in data: data['law_enforcement'] = data['law_enforcement']
    if 'photos' not in data: data['photos'] = []
    
    return data
    
def parse_incident_report(text):
    data = {'raw': text, 'date': text.split('\n')[0], 'status': 'Open', 'photos': []}
    
    fields = [
        'Driver', 'Type', 'Date/Time', 'Location', 'CRS Case #',
        'LMET #', 'Description', 'Statement', 'LMET Called',
        'Law Enforcement', 'Photos'
    ]
    
    # Build pattern that stops at next field name
    for i, field in enumerate(fields):
        next_fields = fields[i+1:] if i+1 < len(fields) else []
        stop = '|'.join([re.escape(f + ':') for f in next_fields]) if next_fields else '$'
        pattern = re.escape(field) + r':\s*(.+?)(?=' + stop + r'|\Z)'
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if match:
            value = match.group(1).strip()
            key = field.lower().replace('/', '_').replace(' ', '_').replace('#', 'number')
            data[key] = value
    
    # Clean up key names
    if 'crs_case_number' in data: data['crs_case'] = data.pop('crs_case_number')
    if 'lmet_number' in data: data['lmet_number'] = data['lmet_number']
    if 'date_time' in data: data['incident_datetime'] = data.pop('date_time')
    if 'lmet_called' in data: data['lmet_called'] = data['lmet_called']
    if 'law_enforcement' in data: data['law_enforcement'] = data['law_enforcement']
    if 'photos' not in data: data['photos'] = []
    
    return data
    
def parse_writeup(text):
    data = {'raw': text, 'date': text.split('\n')[0]}
    employee = re.search(r'Employee:\s*(.+)', text, re.IGNORECASE)
    if employee: data['employee'] = employee.group(1).strip()
    violation = re.search(r'Violation:\s*(.+)', text, re.IGNORECASE)
    if violation: data['violation'] = violation.group(1).strip()
    step = re.search(r'Step:\s*(.+)', text, re.IGNORECASE)
    if step: data['step'] = step.group(1).strip()
    sop = re.search(r'SOP:\s*(.+)', text, re.IGNORECASE)
    if sop: data['sop'] = sop.group(1).strip()
    manager = re.search(r'Manager:\s*(.+)', text, re.IGNORECASE)
    if manager: data['manager'] = manager.group(1).strip()
    description = re.search(r'Description:\s*(.+?)(?=Manager:|Acknowledged|$)', text, re.IGNORECASE | re.DOTALL)
    if description: data['description'] = description.group(1).strip()
    acknowledged = re.search(r'Acknowledged:\s*(.+)', text, re.IGNORECASE)
    if acknowledged: data['acknowledged'] = acknowledged.group(1).strip()
    return data

def parse_fleet_csv(content):
    fleet = []
    reader = csv.DictReader(io.StringIO(content))
    for row in reader:
        van = {
            'id': row.get('vehicleName', '').strip(),
            'status': row.get('status', '').strip(),
            'operational': row.get('operationalStatus', '').strip(),
            'type': row.get('type', '').strip(),
            'ownership': row.get('ownershipType', '').strip(),
            'provider': row.get('vehicleProvider', '').strip(),
            'registration_expiry': row.get('registrationExpiryDate', '').strip(),
            'status_reason': row.get('statusReasonMessage', '').strip(),
            'grounded': 'ground' in row.get('status', '').lower() or 'ground' in row.get('operationalStatus', '').lower(),
            'branded': 'brand' in row.get('ownershipType', '').lower(),
        }
        if van['id']:
            fleet.append(van)
    return fleet

def parse_driver_rating_report(content_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    drivers = []

    def safe_float(val):
        try:
            if val is None or val == 'N/A': return None
            return float(val)
        except: return None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        d = dict(zip(headers, row))
        driver = {
            'name': str(d.get('Driver Name', '')).strip(),
            'transporter_id': str(d.get('Transporter Id', '')).strip(),
            'total_rating': safe_float(d.get('Total Rating')),
            'seatbelt_off': safe_float(d.get('Seatbelt Off Rate')),
            'speeding': safe_float(d.get('Speeding Event Rate')),
            'distractions': safe_float(d.get('Distractions Rate')),
            'following_distance': safe_float(d.get('Following Distance Rate')),
            'sign_signal': safe_float(d.get('Sign/Signal Violations Rate')),
            'cdf_dpmo': safe_float(d.get('Customer Delivery Feedback DPMO')),
            'dc_dpmo': safe_float(d.get('Delivery Completion DPMO')),
            'pod': safe_float(d.get('Photo-On-Delivery')),
            'dsb': safe_float(d.get('DSB Count')),
            'rescue_completed': safe_float(d.get('Rescue Completed')),
            'rescue_requested': safe_float(d.get('Rescue Requested')),
            'callouts': safe_float(d.get('Callouts')),
            'suspensions': safe_float(d.get('Suspensions')),
            'writeups': safe_float(d.get('Write-ups')),
            'damages': safe_float(d.get('Damages')),
            'delivery_success': safe_float(d.get('Delivery Success Behaviors')),
        }
        drivers.append(driver)

    drivers.sort(key=lambda x: x['total_rating'] if x['total_rating'] is not None else 0, reverse=True)
    for i, d in enumerate(drivers):
        d['rank'] = i + 1

    top_rescuers = sorted(
        [d for d in drivers if d['rescue_completed'] and d['rescue_completed'] > 0],
        key=lambda x: x['rescue_completed'], reverse=True
    )[:2]

    flagged = [d for d in drivers if
        (d['writeups'] and d['writeups'] > 0) or
        (d['callouts'] and d['callouts'] > 0) or
        (d['damages'] and d['damages'] > 0)
    ]

    return {
        'drivers': drivers,
        'total_drivers': len(drivers),
        'top_6': drivers[:6],
        'bottom_5': drivers[-5:],
        'top_rescuers': top_rescuers,
        'flagged': flagged
    }

def handle_file(file_info):
    try:
        file_url = file_info.get('url_private_download') or file_info.get('url_private')
        filename = file_info.get('name', '').lower()
        if not file_url:
            return
        headers = {"Authorization": f"Bearer {SLACK_BOT_TOKEN}"}
        response = requests.get(file_url, headers=headers)

        # Save photos from incident reports
        if file_info.get('mimetype', '').startswith('image/'):
            photo_data = {
                'url': file_info.get('permalink', ''),
                'url_private': file_info.get('url_private', ''),
                'name': file_info.get('name', ''),
                'timestamp': file_info.get('timestamp', '')
            }
            existing_photos = redis_get('incident_photos_latest') or []
            existing_photos.append(photo_data)
            redis_set('incident_photos_latest', existing_photos)
            print(f"Photo saved: {photo_data['name']}")
            return

        # LMDmax Driver Rating Report
        if 'driver' in filename and ('rating' in filename or 'report' in filename):
            data = parse_driver_rating_report(response.content)
            redis_set('driver_ratings_latest', data)
            print(f"Driver ratings saved: {data['total_drivers']} drivers")

        # Fleet roster from Amazon AFS
        elif filename.endswith('.xlsx') and ('vehicle' in filename or 'fleet' in filename or 'van' in filename):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(response.content))
            ws = wb.active
            headers_row = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
            fleet = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers_row, row))
                van = {
                    'id': str(row_dict.get('vehicleName', '') or '').strip(),
                    'status': str(row_dict.get('status', '') or '').strip(),
                    'operational': str(row_dict.get('operationalStatus', '') or '').strip(),
                    'type': str(row_dict.get('type', '') or '').strip(),
                    'ownership': str(row_dict.get('ownershipType', '') or '').strip(),
                    'provider': str(row_dict.get('vehicleProvider', '') or '').strip(),
                    'registration_expiry': str(row_dict.get('registrationExpiryDate', '') or '').strip(),
                    'status_reason': str(row_dict.get('statusReasonMessage', '') or '').strip(),
                    'grounded': 'ground' in str(row_dict.get('status', '') or '').lower() or 'ground' in str(row_dict.get('operationalStatus', '') or '').lower(),
                    'branded': 'brand' in str(row_dict.get('ownershipType', '') or '').lower(),
                }
                if van['id'] and van['id'] != 'None':
                    fleet.append(van)
            if fleet:
                redis_set('fleet_roster', fleet)
                print(f"Fleet roster saved: {len(fleet)} vans")

        elif filename.endswith('.csv'):
            fleet = parse_fleet_csv(response.text)
            if fleet:
                redis_set('fleet_roster', fleet)
                print(f"Fleet roster saved: {len(fleet)} vans")

    except Exception as e:
        print(f"File handling error: {e}")

@app.route('/slack/events', methods=['POST'])
def slack_events():
    data = request.json
    if data.get('type') == 'url_verification':
        return jsonify({'challenge': data['challenge']})
    if 'event' in data:
        event = data['event']
        if event.get('type') == 'message' and not event.get('bot_id'):
            text = event.get('text', '')
            print(f"Message received: {text[:200]}")
            if 'Daily Wash Report' in text:
                parsed = parse_daily_wash(text)
                redis_set('daily_wash_latest', parsed)
                print(f"Daily wash saved: {parsed}")
            elif 'Evening Wash Report' in text:
                parsed = parse_evening_wash(text)
                redis_set('evening_wash_latest', parsed)
                print(f"Evening wash saved: {parsed}")
            elif 'Capacity Planning Report' in text:
                parsed = parse_capacity(text)
                redis_set('capacity_latest', parsed)
                print(f"Capacity saved: {parsed}")
            elif 'Fleet Report' in text:
                parsed = parse_fleet_report(text)
                redis_set('fleet_latest', parsed)
                print(f"Fleet saved: {parsed}")
            elif 'Incident Report' in text:
                parsed = parse_incident_report(text)
                existing = redis_get('incidents_all') or []
                existing.append(parsed)
                redis_set('incidents_all', existing)
                print(f"Incident saved: {parsed.get('crs_case', 'No case #')}")
            elif 'Write-Up Report' in text:
                parsed = parse_writeup(text)
                existing = redis_get('writeups_all') or []
                existing.append(parsed)
                redis_set('writeups_all', existing)
                print(f"Write-up saved: {parsed.get('employee', 'Unknown')}")
            elif 'Training Report' in text:
                redis_set('training_latest', {'raw': text, 'date': text.split('\n')[0]})
                print(f"Training saved")
            elif 'Payroll Correction' in text:
                existing = redis_get('payroll_corrections_all') or []
                existing.append({'raw': text, 'date': text.split('\n')[0]})
                redis_set('payroll_corrections_all', existing)
                print(f"Payroll correction saved")
            elif 'Expense Report' in text:
                existing = redis_get('expenses_all') or []
                existing.append({'raw': text, 'date': text.split('\n')[0]})
                redis_set('expenses_all', existing)
                print(f"Expense saved")
            elif 'Hiring Update' in text:
                redis_set('hiring_latest', {'raw': text, 'date': text.split('\n')[0]})
                print(f"Hiring update saved")
            elif 'Termination Report' in text:
                existing = redis_get('terminations_all') or []
                existing.append({'raw': text, 'date': text.split('\n')[0]})
                redis_set('terminations_all', existing)
                print(f"Termination saved")
        if event.get('type') == 'message' and event.get('files'):
            for file_info in event.get('files', []):
                print(f"File received: {file_info.get('name')}")
                handle_file(file_info)
    return jsonify({'status': 'ok'})

@app.route('/data/<key>', methods=['GET'])
def get_data(key):
    headers = {"Authorization": f"Bearer {UPSTASH_TOKEN}"}
    response = requests.get(f"{UPSTASH_URL}/get/{key}", headers=headers)
    result = response.json()
    if result.get('result'):
        try:
            parsed = json.loads(result['result'])
            if isinstance(parsed, str):
                parsed = json.loads(parsed)
            return jsonify({'result': parsed})
        except:
            return jsonify(result)
    return jsonify(result)

@app.route('/test-save', methods=['GET'])
def test_save():
    redis_set('test_key', {'message': 'hello', 'status': 'working'})
    return 'Saved test data to Upstash'

@app.route('/', methods=['GET'])
def home():
    return 'Riptide Command Bot is running.'

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
